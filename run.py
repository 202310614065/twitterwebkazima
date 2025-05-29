import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, simpledialog
import asyncio
import threading
import os
from datetime import datetime, timedelta
from random import randint
from openpyxl import Workbook, load_workbook
from twikit import Client, TooManyRequests, TwitterException
from twikit.errors import BadRequest, NotFound, Forbidden, Unauthorized, AccountLocked
from twikit.tweet import Tweet
import httpx

# Default Configuration (can be overridden by GUI)
DEFAULT_EXCEL_FILE = 'deprem_tweets_gui_output.xlsx'
DEFAULT_START_DT_STR = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d %H:%M:%S')
DEFAULT_END_DT_STR = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
DEFAULT_INTERVAL_HOURS = 1
DEFAULT_TWEETS_PER_INTERVAL = 50 
DEFAULT_SEARCH_PAGE_SIZE = 20 
DEFAULT_REQUEST_DELAY_SEC = 2
DEFAULT_PAGE_REQUEST_DELAY_SEC = 2
DEFAULT_RATE_LIMIT_WAIT_SEC = 5
DEFAULT_LOGIN_RETRY_DELAY_MIN = 10
DEFAULT_LOGIN_RETRY_DELAY_MAX = 20
DEFAULT_MAX_LOGIN_ATTEMPTS_PER_ACCOUNT = 2
DEFAULT_QUERY_KEYWORDS = '("deprem" OR "zelzele")' 
DEFAULT_LANG = 'tr'
DEFAULT_PRODUCT = 'Latest' 

class CriticalClientError(Exception):
    def __init__(self, message, client_identifier="Bilinmeyen Client"):
        super().__init__(message)
        self.client_identifier = client_identifier

class TemporaryClientError(Exception):
    def __init__(self, message, client_identifier="Bilinmeyen Client"):
        super().__init__(message)
        self.client_identifier = client_identifier

class CredentialsDialog(simpledialog.Dialog):
    def __init__(self, parent, title="Twitter Credentials"):
        self.username_var = tk.StringVar()
        self.email_var = tk.StringVar()
        self.password_var = tk.StringVar()
        super().__init__(parent, title)

    def body(self, master):
        ttk.Label(master, text="Twitter Kullanıcı Adı:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
        self.username_entry = ttk.Entry(master, textvariable=self.username_var, width=30)
        self.username_entry.grid(row=0, column=1, padx=5, pady=2)

        ttk.Label(master, text="Twitter E-posta (veya kullanıcı adı):").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
        ttk.Entry(master, textvariable=self.email_var, width=30).grid(row=1, column=1, padx=5, pady=2)

        ttk.Label(master, text="Twitter Şifre:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=2)
        ttk.Entry(master, textvariable=self.password_var, show="*", width=30).grid(row=2, column=1, padx=5, pady=2)
        
        return self.username_entry 

    def apply(self):
        self.result = {
            "username": self.username_var.get(),
            "email": self.email_var.get() or self.username_var.get(), 
            "password": self.password_var.get()
        }

class TwitterClientManager:
    def __init__(self, app_callbacks, lang=DEFAULT_LANG):
        self.app_callbacks = app_callbacks
        self.lang = lang
        self.current_client = None
        self.current_identifier = "N/A"
        self.cookies_file_template = 'cookies_gui_{username}.json'

    async def _login_attempt(self, client, username, email, password, cookie_file):
        login_attempts = 0
        while login_attempts < DEFAULT_MAX_LOGIN_ATTEMPTS_PER_ACCOUNT:
            login_attempts += 1
            self.app_callbacks['log_message'](f"{username} için giriş deneniyor ({login_attempts}/{DEFAULT_MAX_LOGIN_ATTEMPTS_PER_ACCOUNT})...", "INFO")
            try:
                self.app_callbacks['log_message'](f"Eğer Twitter kod isterse, {email} adresine gelen kodu girin.", "USER_INPUT")
                
                auth_info_1 = username
                auth_info_2 = email if email else username
                
                await client.login(
                    auth_info_1=auth_info_1,
                    auth_info_2=auth_info_2,
                    password=password,
                    cookies_file=cookie_file,
                    enable_ui_metrics=True
                )
                self.app_callbacks['log_message'](f"{username} ile giriş yapıldı, {cookie_file} kaydedildi.", "OK")
                return client, username
            except (BadRequest, Forbidden, Unauthorized, AccountLocked) as e:
                self.app_callbacks['log_message'](f"{username} login sırasında API/Hesap Hatası ({type(e).__name__} - {e}).", "ERROR")
                if login_attempts >= DEFAULT_MAX_LOGIN_ATTEMPTS_PER_ACCOUNT:
                    raise ConnectionError(f"{username} ile max login denemesi sonrası başarısız: {e}")
                wait_time = randint(DEFAULT_LOGIN_RETRY_DELAY_MIN * login_attempts, DEFAULT_LOGIN_RETRY_DELAY_MAX * login_attempts)
                self.app_callbacks['log_message'](f"{wait_time} saniye sonra tekrar denenecek...", "INFO")
                await asyncio.sleep(wait_time)
            except TooManyRequests as e:
                self.app_callbacks['log_message'](f"{username} login sırasında Rate Limit ({type(e).__name__} - {e}).", "ERROR")
                if login_attempts >= DEFAULT_MAX_LOGIN_ATTEMPTS_PER_ACCOUNT:
                    raise ConnectionError(f"{username} ile max login denemesi sonrası rate limit: {e}")
                wait_time = randint(DEFAULT_RATE_LIMIT_WAIT_SEC * login_attempts, (DEFAULT_RATE_LIMIT_WAIT_SEC + 10) * login_attempts)
                self.app_callbacks['log_message'](f"Rate limit nedeniyle {wait_time} saniye sonra tekrar denenecek...", "INFO")
                await asyncio.sleep(wait_time)
            except TwitterException as e_twitter:
                self.app_callbacks['log_message'](f"Twitter login hatası ({username}): {e_twitter}.", "ERROR")
                if login_attempts >= DEFAULT_MAX_LOGIN_ATTEMPTS_PER_ACCOUNT:
                    raise ConnectionError(f"{username} ile max deneme sonrası Twitter login hatası: {e_twitter}")
                wait_time = randint(DEFAULT_LOGIN_RETRY_DELAY_MIN * login_attempts, DEFAULT_LOGIN_RETRY_DELAY_MAX * login_attempts)
                self.app_callbacks['log_message'](f"Twitter login hatası nedeniyle {wait_time} saniye sonra tekrar denenecek...", "INFO")
                await asyncio.sleep(wait_time)
            except Exception as e_general:
                self.app_callbacks['log_message'](f"Genel login hatası ({username}): {e_general}.", "ERROR")
                if login_attempts >= DEFAULT_MAX_LOGIN_ATTEMPTS_PER_ACCOUNT:
                    raise ConnectionError(f"{username} ile max deneme sonrası genel login hatası: {e_general}")
                wait_time = randint(DEFAULT_LOGIN_RETRY_DELAY_MIN, DEFAULT_LOGIN_RETRY_DELAY_MAX)
                self.app_callbacks['log_message'](f"Genel login hatası nedeniyle {wait_time} saniye sonra tekrar denenecek...", "INFO")
                await asyncio.sleep(wait_time)
        raise ConnectionError(f"{username} ile tüm giriş denemeleri başarısız.")

    async def ensure_session(self, username, email, password):
        new_client = Client(language=self.lang)
        user_to_log = username if username else "Yeni/Bilinmeyen Hesap"
        cookie_file = self.cookies_file_template.format(username=user_to_log.replace("@","").replace(".","_"))

        if username and os.path.exists(cookie_file):
            try:
                new_client.load_cookies(cookie_file)
                client_username_to_return = user_to_log
                if hasattr(new_client, 'auth_info_1') and new_client.auth_info_1:
                     client_username_to_return = new_client.auth_info_1
                elif hasattr(new_client, 'user') and new_client.user and hasattr(new_client.user, 'username'):
                    client_username_to_return = new_client.user.username

                self.app_callbacks['log_message'](f"{client_username_to_return} (cookie): {cookie_file} yüklendi. Oturum geçerli.", "OK")
                self.current_client = new_client
                self.current_identifier = client_username_to_return
                return True
            except Exception as e:
                self.app_callbacks['log_message'](f"{user_to_log} (cookie): {cookie_file} yüklenirken hata ({type(e).__name__}: {e}), yeniden login.", "WARN")
                if os.path.exists(cookie_file):
                    try:
                        os.remove(cookie_file)
                        self.app_callbacks['log_message'](f"Sorunlu {cookie_file} silindi.", "INFO")
                    except Exception as e_rem:
                        self.app_callbacks['log_message'](f"Cookie ({cookie_file}) silinirken hata: {e_rem}", "ERROR")
        
        if username and password:
            try:
                _, identifier = await self._login_attempt(new_client, username, email, password, cookie_file)
                self.current_client = new_client
                self.current_identifier = identifier
                return True
            except ConnectionError as e:
                self.app_callbacks['log_message'](f"Login failed for {username}: {e}", "ERROR")
                self.current_client = None
                self.current_identifier = "N/A"
                return False
        
        self.app_callbacks['log_message'](f"{user_to_log} için kullanıcı bilgisi sağlanmadı veya login başarısız. Guest moda denenecek.", "WARN")
        try:
            await new_client.login_as_guest()
            self.app_callbacks['log_message'](f"🕵️ {user_to_log} adına Guest moda geçildi.", "OK")
            self.current_client = new_client
            self.current_identifier = "GuestClient"
            return True
        except Exception as e_guest:
            self.app_callbacks['log_message'](f"❌ {user_to_log} adına Guest moda da geçilemedi: {e_guest}", "ERROR")
            self.current_client = None
            self.current_identifier = "N/A"
            return False

    def get_client_details(self):
        return self.current_client, self.current_identifier

class ExcelExporter:
    def __init__(self, filename, app_callbacks):
        self.filename = filename
        self.app_callbacks = app_callbacks
        self.workbook = None
        self.worksheet = None
        self.row_counter = 0
        self._load_or_create_workbook()

    def _load_or_create_workbook(self):
        try:
            self.workbook = load_workbook(self.filename)
            self.worksheet = self.workbook.active
            if self.worksheet.max_row == 0 or (self.worksheet.max_row == 1 and self.worksheet.cell(row=1, column=1).value != '#'):
                self.worksheet.delete_rows(1, self.worksheet.max_row) 
                self._append_header()
            self.row_counter = self.worksheet.max_row
            if self.worksheet.cell(row=1, column=1).value != '#': 
                 self._append_header()
                 self.row_counter = 1

        except FileNotFoundError:
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self._append_header()
            self.row_counter = 1
        except Exception as e:
            self.app_callbacks['log_message'](f"Excel ({self.filename}) yüklenirken hata: {e}. Yeni dosya oluşturuluyor.", "ERROR")
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self._append_header()
            self.row_counter = 1
        self.app_callbacks['update_excel_tweets_count'](self.row_counter -1 if self.row_counter > 0 else 0)


    def _append_header(self):
        self.worksheet.append(['#', 'Kullanıcı', 'Tarih', 'Tweet', 'RT', 'Likes'])

    def append_tweets(self, tweets_data):
        if not self.worksheet:
            return
        for t_data in tweets_data:
            self.row_counter += 1
            self.worksheet.append([
                self.row_counter -1,
                t_data.get('user_name', 'N/A'),
                t_data.get('date_str', 'N/A'),
                t_data.get('text', ''),
                t_data.get('retweet_count', 0),
                t_data.get('favorite_count', 0)
            ])
        self.app_callbacks['update_excel_tweets_count'](self.row_counter -1 if self.row_counter > 0 else 0)


    def save_workbook(self):
        if not self.workbook:
            return False
        try:
            self.workbook.save(self.filename)
            self.app_callbacks['log_message'](f"Excel dosyası '{self.filename}' kaydedildi.", "OK")
            return True
        except PermissionError:
            new_filename = self.filename.replace(".xlsx", f"_locked_{datetime.now():%H%M%S}.xlsx")
            try:
                self.workbook.save(new_filename)
                self.app_callbacks['log_message'](f"'{self.filename}' kilitli. '{new_filename}' olarak kaydedildi.", "ERROR")
                self.filename = new_filename 
                return True
            except Exception as e_new:
                self.app_callbacks['log_message'](f"Excel'i yeni isimle ({new_filename}) kaydetme hatası: {e_new}", "ERROR")
                return False
        except Exception as e:
            self.app_callbacks['log_message'](f"Excel kaydetme hatası: {e}", "ERROR")
            return False

class TwitterScraper:
    def __init__(self, app_callbacks, query_params):
        self.app_callbacks = app_callbacks
        self.query_params = query_params
        self.client_manager = TwitterClientManager(app_callbacks, query_params.get('lang', DEFAULT_LANG))
        self.excel_exporter = ExcelExporter(query_params.get('excel_file', DEFAULT_EXCEL_FILE), app_callbacks)
        
        self.is_running = False
        self.is_paused = False
        self.stop_requested = False
        self.current_task_state = None 
        self.collected_tweet_ids_total_run = set() 
        self.client_ready_event = asyncio.Event()
        self.loop = None

    async def _initialize_client(self, credentials):
        success = await self.client_manager.ensure_session(
            credentials['username'], credentials['email'], credentials['password']
        )
        if success:
            _, identifier = self.client_manager.get_client_details()
            self.app_callbacks['update_current_account'](identifier)
            self.client_ready_event.set()
            return True
        else:
            self.app_callbacks['update_current_account']("Giriş Başarısız")
            self.client_ready_event.clear() 
            return False

    def start_scraping_thread(self, initial_credentials):
        self.is_running = True
        self.is_paused = False
        self.stop_requested = False
        self.collected_tweet_ids_total_run.clear()
        
        self.loop = asyncio.new_event_loop()
        asyncio.set_event_loop(self.loop)
        
        self.loop.run_until_complete(self._initialize_client(initial_credentials))
        if not self.client_ready_event.is_set():
            self.app_callbacks['log_message']("Başlangıç client oluşturulamadı. Scraping başlatılamıyor.", "CRITICAL")
            self.is_running = False
            self.app_callbacks['on_scraping_finished']() 
            return

        self.loop.create_task(self._scraping_loop())
        
        def run_loop():
            try:
                self.loop.run_forever()
            finally:
                self.loop.close()
        
        self.thread = threading.Thread(target=run_loop, daemon=True)
        self.thread.start()

    def pause_scraping(self):
        if self.is_running and not self.is_paused:
            self.is_paused = True
            self.app_callbacks['log_message']("Scraping duraklatıldı.", "INFO")
            self.app_callbacks['update_status']("Duraklatıldı")

    def resume_scraping(self):
        if self.is_running and self.is_paused:
            if not self.client_ready_event.is_set():
                self.app_callbacks['log_message']("Client hazır değil. Devam ettirmeden önce giriş yapın/hesap değiştirin.", "WARN")
                self.app_callbacks['request_new_credentials_for_resume'](self.current_task_state)
                return
            self.is_paused = False
            self.app_callbacks['log_message']("Scraping devam ediyor...", "INFO")

    def stop_scraping(self):
        if self.is_running:
            self.stop_requested = True
            self.is_running = False 
            self.is_paused = False 
            self.app_callbacks['log_message']("Scraping durduruluyor...", "INFO")
            if self.loop and not self.loop.is_closed():
                 self.loop.call_soon_threadsafe(self.loop.stop)


    def save_current_data(self):
        self.app_callbacks['log_message']("Mevcut veriler kaydediliyor...", "INFO")
        if self.excel_exporter.save_workbook():
            self.app_callbacks['log_message']("Veriler başarıyla kaydedildi.", "OK")
        else:
            self.app_callbacks['log_message']("Veri kaydetme başarısız.", "ERROR")

    async def switch_account_and_resume(self, new_credentials, resume_state):
        self.current_task_state = resume_state 
        self.is_paused = True 
        self.client_ready_event.clear()
        
        self.app_callbacks['log_message'](f"Yeni hesap ({new_credentials.get('username', 'Bilinmeyen')}) ile devam edilecek...", "INFO")
        
        login_success = await self._initialize_client(new_credentials)
        if login_success:
            self.is_paused = False 
            self.app_callbacks['log_message']("Hesap değiştirildi, scraping devam edecek.", "OK")
        else:
            self.app_callbacks['log_message']("Yeni hesapla giriş başarısız. Scraping duraklatıldı.", "ERROR")

    def _build_query(self, since: datetime, until: datetime, keywords, lang, max_id=None):
        s_utc = since.strftime('%Y-%m-%d_%H:%M:%S_UTC')
        u_utc = until.strftime('%Y-%m-%d_%H:%M:%S_UTC')
        query = f'{keywords} since:{s_utc} until:{u_utc} lang:{lang}'
        if max_id:
            query += f' max_id:{max_id}'
        return query

    async def _fetch_page_data(self, client, query, product, count, client_identifier, since_dt, until_dt):
        page_retries = 0
        max_retries_default = 3
        max_retries_ratelimit = 2 

        while page_retries < max_retries_default:
            try:
                self.app_callbacks['log_message'](f"{client_identifier} | {since_dt.strftime('%H:%M')}–{until_dt.strftime('%H:%M')}: search_tweet (Query: '{query[:70]}...'), Deneme: {page_retries+1}", "DEBUG")
                raw_page_results = await client.search_tweet(query=query, product=product, count=count)
                return raw_page_results
            except (Forbidden, Unauthorized, AccountLocked) as e:
                self.app_callbacks['log_message'](f"API Yetki/Hesap Kilit Hatası ({type(e).__name__}: {e}) | {client_identifier}.", "ERROR")
                raise CriticalClientError(f"Client yetkisi sonlandı/hesap kilitli: {e}", client_identifier)
            except TooManyRequests as e:
                if page_retries < max_retries_ratelimit:
                    page_retries += 1
                    self.app_callbacks['log_message'](f"search_tweet Rate limit | {client_identifier}. {DEFAULT_RATE_LIMIT_WAIT_SEC}s uyku (Deneme {page_retries}/{max_retries_ratelimit+1})...", "WARN")
                    await asyncio.sleep(DEFAULT_RATE_LIMIT_WAIT_SEC)
                else:
                    self.app_callbacks['log_message'](f"{client_identifier} | Sürekli rate-limit. Client değiştirme sinyali.", "ERROR")
                    raise TemporaryClientError(f"Rate limit aşıldı: {e}", client_identifier) 
            except TwitterException as e:
                err_msg_lower = str(e).lower()
                if any(keyword in err_msg_lower for keyword in ["suspended", "terminated", "deactivated", "restricted"]):
                    self.app_callbacks['log_message'](f"Kritik Twitter Hesap Hatası ({type(e).__name__}: {e}) | {client_identifier}.", "ERROR")
                    raise CriticalClientError(f"Kritik Twitter Hesap Hatası: {e}", client_identifier)
                
                self.app_callbacks['log_message'](f"search_tweet sırasında TwitterException ({e}) | {client_identifier}. Sayfa atlanıyor.", "WARN")
                return [] 
            except (httpx.ConnectTimeout, httpx.ReadTimeout, httpx.ConnectError, httpx.NetworkError) as e:
                page_retries +=1
                wait_net = randint(10,20)
                self.app_callbacks['log_message'](f"Ağ hatası ({type(e).__name__}) | {client_identifier}. {wait_net}s uyku (Deneme {page_retries}/{max_retries_default})...", "WARN")
                await asyncio.sleep(wait_net)
                if page_retries >= max_retries_default:
                    self.app_callbacks['log_message'](f"{client_identifier} | Sürekli ağ hatası. Sayfa atlanıyor.", "ERROR")
                    return []
            except Exception as e:
                self.app_callbacks['log_message'](f"Genel hata ({type(e).__name__}: {e}) | {client_identifier}. Sayfa atlanıyor.", "ERROR")
                return [] 
        return [] 


    async def _fetch_interval_data(self, since_dt, until_dt):
        client, client_identifier = self.client_manager.get_client_details()
        if not client:
            raise CriticalClientError("Client not available for fetching interval.", "N/A")

        keywords = self.query_params.get('keywords', DEFAULT_QUERY_KEYWORDS)
        lang = self.query_params.get('lang', DEFAULT_LANG)
        product = self.query_params.get('product', DEFAULT_PRODUCT)
        tweets_per_interval_target = self.query_params.get('tweets_per_interval', DEFAULT_TWEETS_PER_INTERVAL)
        search_page_size = self.query_params.get('search_page_size', DEFAULT_SEARCH_PAGE_SIZE)
        page_request_delay = self.query_params.get('page_request_delay_sec', DEFAULT_PAGE_REQUEST_DELAY_SEC)

        current_max_id = self.current_task_state.get('max_id')
        page_num_start = self.current_task_state.get('page_num', 0)
        interval_tweets_collected_count = self.current_task_state.get('collected_in_interval', 0)
        
        interval_tweets_data = []
        collected_tweet_ids_this_interval = set() 

        max_page_fetches = (tweets_per_interval_target // (search_page_size // 2 if search_page_size > 1 else 1) ) + 10

        for page_num in range(page_num_start, max_page_fetches):
            if self.stop_requested or not self.is_running: break
            while self.is_paused: await asyncio.sleep(0.1)
            if self.stop_requested or not self.is_running: break

            self.current_task_state['page_num'] = page_num
            query_str = self._build_query(since_dt, until_dt, keywords, lang, current_max_id)
            raw_page_results = await self._fetch_page_data(client, query_str, product, search_page_size, client_identifier, since_dt, until_dt)
            
            current_page_new_tweets_count = 0
            if raw_page_results:
                new_tweets_on_page = []
                for item in raw_page_results:
                    if isinstance(item, Tweet) and item.id not in self.collected_tweet_ids_total_run and item.id not in collected_tweet_ids_this_interval:
                        new_tweets_on_page.append(item)
                        collected_tweet_ids_this_interval.add(item.id)
                        self.collected_tweet_ids_total_run.add(item.id)
                
                if new_tweets_on_page:
                    for t in new_tweets_on_page:
                        date_str = str(getattr(t, 'created_at', 'N/A'))
                        try:
                            dt_obj = t.created_at if isinstance(t.created_at, datetime) else datetime.strptime(str(t.created_at).split('.')[0], '%Y-%m-%dT%H:%M:%S')
                            date_str = dt_obj.strftime('%Y-%m-%d %H:%M:%S')
                        except: pass
                        
                        tweet_data = {
                            'id': t.id,
                            'user_name': getattr(getattr(t, 'user', None), 'name', 'N/A'),
                            'date_str': date_str,
                            'text': getattr(t, 'text', '').replace('\n',' ').replace('\r',''),
                            'retweet_count': getattr(t, 'retweet_count', 0),
                            'favorite_count': getattr(t, 'favorite_count', 0)
                        }
                        interval_tweets_data.append(tweet_data)
                        current_page_new_tweets_count +=1
                    
                    interval_tweets_collected_count += current_page_new_tweets_count
                    self.current_task_state['collected_in_interval'] = interval_tweets_collected_count
                    self.app_callbacks['log_message'](f"{client_identifier} | {since_dt.strftime('%H:%M')}–{until_dt.strftime('%H:%M')}: Sayfa {page_num+1} - {current_page_new_tweets_count} yeni. Aralıkta: {interval_tweets_collected_count}/{tweets_per_interval_target}", "OK")
                    
                    oldest_tweet_in_page = new_tweets_on_page[-1]
                    current_max_id = str(int(oldest_tweet_in_page.id) - 1)
                    self.current_task_state['max_id'] = current_max_id
                else: 
                    self.app_callbacks['log_message'](f"{client_identifier} | {since_dt.strftime('%H:%M')}–{until_dt.strftime('%H:%M')}: Sayfa {page_num+1}'da yeni tweet yok.", "INFO")
                    break 
            else: 
                self.app_callbacks['log_message'](f"{client_identifier} | {since_dt.strftime('%H:%M')}–{until_dt.strftime('%H:%M')}: Sayfa {page_num+1}'dan sonuç alınamadı.", "INFO")
                break 

            if interval_tweets_collected_count >= tweets_per_interval_target:
                self.app_callbacks['log_message'](f"{client_identifier} | {since_dt.strftime('%H:%M')}–{until_dt.strftime('%H:%M')}: Hedef {tweets_per_interval_target} tweete ulaşıldı.", "INFO")
                break
            
            if current_page_new_tweets_count > 0 : 
                 await asyncio.sleep(page_request_delay)

        self.app_callbacks['log_message'](f"{client_identifier} | {since_dt.strftime('%Y-%m-%d %H:%M')}–{until_dt.strftime('%Y-%m-%d %H:%M')}: Toplam {interval_tweets_collected_count} tweet çekildi.", "INFO")
        return interval_tweets_data


    async def _scraping_loop(self):
        start_dt_str = self.query_params.get('start_dt', DEFAULT_START_DT_STR)
        end_dt_str = self.query_params.get('end_dt', DEFAULT_END_DT_STR)
        interval_hours = self.query_params.get('interval_hours', DEFAULT_INTERVAL_HOURS)
        request_delay_sec = self.query_params.get('request_delay_sec', DEFAULT_REQUEST_DELAY_SEC)

        try:
            start_dt = datetime.strptime(start_dt_str, '%Y-%m-%d %H:%M:%S')
            end_dt = datetime.strptime(end_dt_str, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            self.app_callbacks['log_message']("Geçersiz tarih formatı. YYYY-MM-DD HH:MM:SS kullanın.", "CRITICAL")
            self.is_running = False
            self.app_callbacks['on_scraping_finished']()
            return

        current_dt = start_dt
        if self.current_task_state and 'since' in self.current_task_state and isinstance(self.current_task_state['since'], datetime): 
            current_dt = self.current_task_state['since']
            self.app_callbacks['log_message'](f"Scraping {current_dt.strftime('%Y-%m-%d %H:%M:%S')} tarihinden devam ediyor...", "INFO")
        else: 
             self.current_task_state = {'since': current_dt}


        total_intervals_approx = max(1, (end_dt - start_dt).total_seconds() / (interval_hours * 3600))
        interval_num_display = 0 

        temp_dt_for_interval_count = start_dt
        while temp_dt_for_interval_count < current_dt:
            interval_num_display +=1
            temp_dt_for_interval_count += timedelta(hours=interval_hours)


        while current_dt < end_dt and self.is_running and not self.stop_requested:
            interval_num_display +=1
            while self.is_paused and self.is_running and not self.stop_requested:
                await asyncio.sleep(0.5)
            
            if self.stop_requested or not self.is_running: break

            if not self.client_ready_event.is_set():
                self.app_callbacks['log_message']("Client hazır değil, bekleniyor...", "WARN")
                self.is_paused = True 
                self.app_callbacks['update_status'](f"Hesap bekleniyor...")
                await self.client_ready_event.wait() 
                self.is_paused = False 
                self.app_callbacks['log_message']("Client hazır, devam ediliyor.", "INFO")

            since_dt = current_dt
            until_dt = min(current_dt + timedelta(hours=interval_hours), end_dt)
            
            if not (self.current_task_state and self.current_task_state.get('since') == since_dt and self.current_task_state.get('until') == until_dt) :
                 self.current_task_state = {'since': since_dt, 'until': until_dt, 'max_id': None, 'page_num': 0, 'collected_in_interval': 0}

            self.app_callbacks['update_status'](f"Aralık {interval_num_display}/{int(total_intervals_approx)}: {since_dt.strftime('%y-%m-%d %H:%M')} - {until_dt.strftime('%y-%m-%d %H:%M')}")
            
            try:
                interval_tweets = await self._fetch_interval_data(since_dt, until_dt)
                if interval_tweets:
                    self.excel_exporter.append_tweets(interval_tweets)
                
                current_dt = until_dt
                self.current_task_state = {'since': current_dt} 
                if current_dt < end_dt and self.is_running: 
                    await asyncio.sleep(request_delay_sec)

            except (CriticalClientError, TemporaryClientError) as e:
                self.app_callbacks['log_message'](f"Client hatası: {e}. Yeni hesap bilgileri gerekiyor.", "ERROR")
                self.is_paused = True
                self.client_ready_event.clear()
                self.app_callbacks['request_new_credentials_for_resume'](self.current_task_state)
            except Exception as e:
                self.app_callbacks['log_message'](f"Aralık işlenirken genel hata ({type(e).__name__}: {e}). Bu aralık atlanıyor.", "ERROR")
                current_dt = until_dt 
                self.current_task_state = {'since': current_dt} 
                if current_dt < end_dt and self.is_running:
                    await asyncio.sleep(request_delay_sec) 

        if self.is_running and not self.stop_requested:
            self.app_callbacks['log_message']("Tüm aralıklar tamamlandı.", "OK")
            self.app_callbacks['update_status']("Tamamlandı")
        elif self.stop_requested:
            self.app_callbacks['log_message']("Scraping kullanıcı tarafından durduruldu.", "INFO")
            self.app_callbacks['update_status']("Durduruldu")
        
        self.is_running = False
        self.excel_exporter.save_workbook()
        self.app_callbacks['on_scraping_finished']()


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Twitter Scraper GUI")
        self.geometry("1000x750") 

        self.query_params = {}
        self.scraper = None
        self.credentials_dialog_open = False

        self.callbacks = {
            'log_message': self.log_message,
            'update_status': self.update_status,
            'update_current_account': self.update_current_account,
            'update_excel_tweets_count': self.update_excel_tweets_count,
            'request_new_credentials_for_resume': self.handle_request_new_credentials,
            'on_scraping_finished': self.on_scraping_operation_finished,
        }
        
        self.constructed_keywords_var = tk.StringVar(value=DEFAULT_QUERY_KEYWORDS) 
        self.keyword_entry_var = tk.StringVar()
        self.min_retweets_var = tk.StringVar(value="0")
        self.min_faves_var = tk.StringVar(value="0")

        self.init_query_builder_ui()

    def log_message(self, msg, level="INFO"):
        if hasattr(self, 'log_text_widget'): 
            log_time = datetime.now().strftime('%H:%M:%S')
            formatted_msg = f"{log_time} | {level:<7} | {msg}\n"
            self.log_text_widget.insert(tk.END, formatted_msg)
            self.log_text_widget.see(tk.END)
            if level in ["ERROR", "CRITICAL", "WARN"]:
                 print(formatted_msg.strip()) 

    def update_status(self, status_msg):
        if hasattr(self, 'status_label'):
            self.status_label.config(text=f"Durum: {status_msg}")

    def update_current_account(self, account_name):
        if hasattr(self, 'account_label'):
            self.account_label.config(text=f"Aktif Hesap: {account_name}")
    
    def update_excel_tweets_count(self, count):
        if hasattr(self, 'tweets_collected_label'):
            self.tweets_collected_label.config(text=f"Excel'deki Tweet Sayısı: {count}")

    def _validate_entry(self, P, max_length_str):
        max_length = int(max_length_str)
        if P == "": return True 
        if not P.isdigit(): return False
        if len(P) > max_length: return False
        return True

    def _add_to_keyword_query(self, text_to_add, needs_space_before=True, needs_space_after=True):
        current_query = self.constructed_keywords_var.get()
        new_text = ""
        if current_query and needs_space_before and not current_query.endswith(" ") and not text_to_add.startswith(" "):
            new_text += " "
        
        new_text += text_to_add

        if needs_space_after and not text_to_add.endswith(" "):
             new_text += " "
        
        # Avoid leading operators if query is empty
        if not current_query and text_to_add.strip() in ["OR", "AND"]:
            return

        self.constructed_keywords_var.set(current_query + new_text.rstrip() if text_to_add.endswith(")") else current_query + new_text)


    def _add_keyword_from_entry(self):
        keyword = self.keyword_entry_var.get().strip()
        if keyword:
            if " " in keyword and not (keyword.startswith('"') and keyword.endswith('"')): 
                keyword = f'"{keyword}"'
            self._add_to_keyword_query(keyword, needs_space_after=True)
            self.keyword_entry_var.set("") 

    def _clear_keyword_query(self):
        self.constructed_keywords_var.set("")

    def _add_filter_keyword(self, filter_text):
        self._add_to_keyword_query(filter_text, needs_space_after=True)

    def init_query_builder_ui(self):
        self.query_frame = ttk.Frame(self, padding="10")
        self.query_frame.pack(expand=True, fill=tk.BOTH)
        
        ttk.Label(self.query_frame, text="Twitter Web Kazıma | github.com/202310614065", font=("Arial", 16)).pack(pady=10)

        main_param_scroll_canvas = tk.Canvas(self.query_frame)
        scrollbar = ttk.Scrollbar(self.query_frame, orient="vertical", command=main_param_scroll_canvas.yview)
        scrollable_frame = ttk.Frame(main_param_scroll_canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: main_param_scroll_canvas.configure(
                scrollregion=main_param_scroll_canvas.bbox("all")
            )
        )
        main_param_scroll_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        main_param_scroll_canvas.configure(yscrollcommand=scrollbar.set)

        main_param_scroll_canvas.pack(side="left", fill="both", expand=True, pady=5)
        scrollbar.pack(side="right", fill="y")

        kw_frame = ttk.LabelFrame(scrollable_frame, text="Anahtar Kelime ve Operatörler", padding="10")
        kw_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Entry(kw_frame, textvariable=self.keyword_entry_var, width=40).grid(row=0, column=0, columnspan=3, padx=2, pady=2, sticky=tk.EW)
        ttk.Button(kw_frame, text="Kelime Ekle", command=self._add_keyword_from_entry).grid(row=0, column=3, padx=2, pady=2)
        
        op_buttons_frame = ttk.Frame(kw_frame)
        op_buttons_frame.grid(row=1, column=0, columnspan=4, pady=5)
        ttk.Button(op_buttons_frame, text='( Aç', command=lambda: self._add_to_keyword_query("(", False, False)).pack(side=tk.LEFT, padx=2)
        ttk.Button(op_buttons_frame, text=') Kapat', command=lambda: self._add_to_keyword_query(")", True, True)).pack(side=tk.LEFT, padx=2)
        ttk.Button(op_buttons_frame, text="OR", command=lambda: self._add_to_keyword_query("OR", True, True)).pack(side=tk.LEFT, padx=2)
        ttk.Button(op_buttons_frame, text="AND", command=lambda: self._add_to_keyword_query("AND", True, True)).pack(side=tk.LEFT, padx=2)
        ttk.Button(op_buttons_frame, text="Temizle", command=self._clear_keyword_query).pack(side=tk.LEFT, padx=10)

        ttk.Label(kw_frame, text="Oluşturulan Sorgu (Düzenlenebilir):").grid(row=2, column=0, columnspan=4, sticky=tk.W, pady=(5,0))
        kw_display = ttk.Entry(kw_frame, textvariable=self.constructed_keywords_var, width=60, state='normal') # Editable
        kw_display.grid(row=3, column=0, columnspan=4, sticky=tk.EW, padx=2, pady=2)
        
        filter_frame = ttk.LabelFrame(scrollable_frame, text="Filtreler", padding="10")
        filter_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Label(filter_frame, text="Min Retweet:").grid(row=0, column=0, padx=2, pady=2, sticky=tk.W)
        ttk.Entry(filter_frame, textvariable=self.min_retweets_var, width=5).grid(row=0, column=1, padx=2, pady=2)
        ttk.Button(filter_frame, text="Ekle", command=lambda: self._add_filter_keyword(f"min_retweets:{self.min_retweets_var.get()}")).grid(row=0, column=2, padx=2, pady=2)

        ttk.Label(filter_frame, text="Min Beğeni:").grid(row=1, column=0, padx=2, pady=2, sticky=tk.W)
        ttk.Entry(filter_frame, textvariable=self.min_faves_var, width=5).grid(row=1, column=1, padx=2, pady=2)
        ttk.Button(filter_frame, text="Ekle", command=lambda: self._add_filter_keyword(f"min_faves:{self.min_faves_var.get()}")).grid(row=1, column=2, padx=2, pady=2)
        
        filter_buttons_frame = ttk.Frame(filter_frame)
        filter_buttons_frame.grid(row=2, column=0, columnspan=3, pady=5)
        ttk.Button(filter_buttons_frame, text="Doğrulanmış Hesaplar", command=lambda: self._add_filter_keyword("filter:verified")).pack(side=tk.LEFT, padx=2)
        ttk.Button(filter_buttons_frame, text="Güvenli Tweetler", command=lambda: self._add_filter_keyword("filter:safe")).pack(side=tk.LEFT, padx=2)
        ttk.Button(filter_buttons_frame, text="Yanıtları Hariç Tut", command=lambda: self._add_filter_keyword("-filter:replies")).pack(side=tk.LEFT, padx=2)

        self.q_params_vars = {} 
        
        def create_datetime_entries(parent, label_text, default_dt_str, prefix):
            frame = ttk.LabelFrame(parent, text=label_text, padding="5")
            frame.pack(fill=tk.X, padx=5, pady=5)
            
            dt_obj = datetime.strptime(default_dt_str, '%Y-%m-%d %H:%M:%S')
            
            parts = {"Y": dt_obj.year, "M": dt_obj.month, "D": dt_obj.day, "h": dt_obj.hour, "m": dt_obj.minute, "s": dt_obj.second}
            widths = {"Y": 4, "M": 2, "D": 2, "h": 2, "m": 2, "s": 2}
            labels = {"Y": "Yıl", "M": "Ay", "D": "Gün", "h": "Sa", "m": "Dk", "s": "Sn"}

            col = 0
            for key, default_val in parts.items():
                ttk.Label(frame, text=labels[key] + ":").grid(row=0, column=col, padx=(5,0), pady=2)
                col+=1
                var = tk.StringVar(value=str(default_val).zfill(widths[key]))
                self.q_params_vars[f"{prefix}_{key}"] = var
                vcmd = (self.register(self._validate_entry), '%P', str(widths[key]))
                ttk.Entry(frame, textvariable=var, width=widths[key]+1, validate='key', validatecommand=vcmd).grid(row=0, column=col, padx=(0,5), pady=2)
                col+=1
            return frame

        create_datetime_entries(scrollable_frame, "Başlangıç Tarihi ve Saati", DEFAULT_START_DT_STR, "start_dt")
        create_datetime_entries(scrollable_frame, "Bitiş Tarihi ve Saati", DEFAULT_END_DT_STR, "end_dt")

        other_params_frame = ttk.LabelFrame(scrollable_frame, text="Diğer Ayarlar", padding="10")
        other_params_frame.pack(fill=tk.X, padx=5, pady=5)
        
        other_params_config = [
            ("Dil Kodu", "lang", DEFAULT_LANG, 10),
            ("Tweet Türü (Latest/Top)", "product", DEFAULT_PRODUCT, 15, ["Latest", "Top"]), 
            ("Aralık Başına Tweet Sayısı", "tweets_per_interval", str(DEFAULT_TWEETS_PER_INTERVAL), 5),
            ("Sayfa Başına Tweet (API)", "search_page_size", str(DEFAULT_SEARCH_PAGE_SIZE), 5),
            ("Aralık Adımı (saat)", "interval_hours", str(DEFAULT_INTERVAL_HOURS), 5),
            ("Excel Dosya Adı", "excel_file", DEFAULT_EXCEL_FILE, 30),
            ("İstekler Arası Gecikme (sn)", "request_delay_sec", str(DEFAULT_REQUEST_DELAY_SEC), 5),
            ("Sayfa İstekleri Arası Gecikme (sn)", "page_request_delay_sec", str(DEFAULT_PAGE_REQUEST_DELAY_SEC), 5),
        ]

        for i, item in enumerate(other_params_config):
            label_text, key, default_val, width = item[0], item[1], item[2], item[3]
            options = item[4] if len(item) > 4 else None

            ttk.Label(other_params_frame, text=label_text + ":").grid(row=i, column=0, sticky=tk.W, padx=5, pady=2)
            var = tk.StringVar(value=default_val)
            self.q_params_vars[key] = var
            if options: 
                ttk.Combobox(other_params_frame, textvariable=var, values=options, width=width-2, state="readonly").grid(row=i, column=1, sticky=tk.EW, padx=5, pady=2)
            else: 
                ttk.Entry(other_params_frame, textvariable=var, width=width).grid(row=i, column=1, sticky=tk.EW, padx=5, pady=2)
        
        other_params_frame.grid_columnconfigure(1, weight=1)
        
        ttk.Button(self.query_frame, text="Yapılandır ve Başlat", command=self.start_scraping_with_params).pack(pady=15, side=tk.BOTTOM)


    def start_scraping_with_params(self):
        self.query_params = {} 
        
        keywords_query = self.constructed_keywords_var.get().strip()
        if not keywords_query:
            messagebox.showerror("Parametre Hatası", "Lütfen anahtar kelime sorgusu oluşturun.")
            return
        self.query_params['keywords'] = keywords_query

        try:
            start_dt_str_parts = {}
            for part_key in ["Y", "M", "D", "h", "m", "s"]:
                val = self.q_params_vars[f"start_dt_{part_key}"].get()
                if not val or not val.isdigit(): 
                    raise ValueError(f"Başlangıç tarihi '{part_key}' kısmı sayısal değil veya boş: '{val}'")
                start_dt_str_parts[part_key] = val
            self.query_params['start_dt'] = f"{start_dt_str_parts['Y']}-{start_dt_str_parts['M']}-{start_dt_str_parts['D']} {start_dt_str_parts['h']}:{start_dt_str_parts['m']}:{start_dt_str_parts['s']}"
            datetime.strptime(self.query_params['start_dt'], '%Y-%m-%d %H:%M:%S')

            end_dt_str_parts = {}
            for part_key in ["Y", "M", "D", "h", "m", "s"]:
                val = self.q_params_vars[f"end_dt_{part_key}"].get()
                if not val or not val.isdigit():
                    raise ValueError(f"Bitiş tarihi '{part_key}' kısmı sayısal değil veya boş: '{val}'")
                end_dt_str_parts[part_key] = val
            self.query_params['end_dt'] = f"{end_dt_str_parts['Y']}-{end_dt_str_parts['M']}-{end_dt_str_parts['D']} {end_dt_str_parts['h']}:{end_dt_str_parts['m']}:{end_dt_str_parts['s']}"
            datetime.strptime(self.query_params['end_dt'], '%Y-%m-%d %H:%M:%S')

        except ValueError as e_val: 
            messagebox.showerror("Tarih/Saat Hatası", f"Lütfen tüm tarih ve saat alanlarını doğru ve sayısal olarak doldurun.\nDetay: {e_val}")
            return
        except KeyError as e_key: 
            messagebox.showerror("Program Hatası", f"Tarih parametresi anahtarı bulunamadı: {e_key}")
            return

        try:
            for key in ["lang", "product", "excel_file"]: 
                 self.query_params[key] = self.q_params_vars[key].get()

            for key in ["tweets_per_interval", "search_page_size", "interval_hours", 
                        "request_delay_sec", "page_request_delay_sec"]: 
                self.query_params[key] = int(self.q_params_vars[key].get())
        except ValueError as e:
            messagebox.showerror("Parametre Hatası", f"Lütfen sayısal alanları doğru formatta girin.\nHata: {e}")
            return
        except KeyError as e:
            messagebox.showerror("Parametre Hatası", f"Eksik parametre yapılandırması: {e}")
            return

        self.query_frame.pack_forget()
        self.init_main_app_ui()
        self.prompt_initial_credentials()
        
    def prompt_initial_credentials(self):
        if self.credentials_dialog_open: return
        self.credentials_dialog_open = True
        dialog = CredentialsDialog(self, title="Başlangıç Twitter Hesabı")
        self.credentials_dialog_open = False 
        if dialog.result:
            creds = dialog.result
            if not creds['username'] or not creds['password']:
                messagebox.showerror("Eksik Bilgi", "Kullanıcı adı ve şifre gereklidir.")
                if hasattr(self, 'main_app_frame') and self.main_app_frame.winfo_exists():
                    self.main_app_frame.pack_forget()
                self.init_query_builder_ui() 
                return
            
            self.scraper = TwitterScraper(self.callbacks, self.query_params)
            self.scraper.start_scraping_thread(creds)
            self.update_gui_for_scraping_active(True)
        else: 
            messagebox.showinfo("İptal Edildi", "Scraping başlatılmadı.")
            if hasattr(self, 'main_app_frame') and self.main_app_frame.winfo_exists():
                 self.main_app_frame.pack_forget()
            self.init_query_builder_ui()


    def init_main_app_ui(self):
        self.main_app_frame = ttk.Frame(self, padding="5")
        self.main_app_frame.pack(expand=True, fill=tk.BOTH)

        control_panel = ttk.Frame(self.main_app_frame, width=250, relief=tk.RIDGE, padding="5")
        control_panel.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)
        control_panel.pack_propagate(False)

        self.account_label = ttk.Label(control_panel, text="Aktif Hesap: N/A", font=("Arial", 10))
        self.account_label.pack(pady=5, anchor=tk.W)
        self.status_label = ttk.Label(control_panel, text="Durum: Beklemede", font=("Arial", 10))
        self.status_label.pack(pady=5, anchor=tk.W)
        self.tweets_collected_label = ttk.Label(control_panel, text="Excel'deki Tweet Sayısı: 0", font=("Arial", 10))
        self.tweets_collected_label.pack(pady=5, anchor=tk.W)

        self.pause_resume_button = ttk.Button(control_panel, text="Duraklat", command=self.toggle_pause_resume)
        self.pause_resume_button.pack(pady=5, fill=tk.X)
        
        self.switch_account_button = ttk.Button(control_panel, text="Hesap Değiştir", command=self.handle_switch_account_button)
        self.switch_account_button.pack(pady=5, fill=tk.X)

        self.save_button = ttk.Button(control_panel, text="Mevcut Veriyi Kaydet", command=self.handle_save_button)
        self.save_button.pack(pady=5, fill=tk.X)
        
        self.stop_button = ttk.Button(control_panel, text="Durdur ve Çık", command=self.handle_stop_button)
        self.stop_button.pack(pady=10, side=tk.BOTTOM, fill=tk.X)

        log_panel = ttk.Frame(self.main_app_frame, relief=tk.RIDGE, padding="5")
        log_panel.pack(side=tk.RIGHT, expand=True, fill=tk.BOTH, padx=5, pady=5)
        
        ttk.Label(log_panel, text="İşlem Kayıtları", font=("Arial", 12)).pack(pady=5)
        self.log_text_widget = scrolledtext.ScrolledText(log_panel, wrap=tk.WORD, height=10, width=70, font=("Courier New", 9))
        self.log_text_widget.pack(expand=True, fill=tk.BOTH)
        self.update_gui_for_scraping_active(False) 

    def handle_request_new_credentials(self, resume_state_info):
        if self.credentials_dialog_open:
            self.log_message("Credential dialog zaten açık.", "WARN")
            return
        self.credentials_dialog_open = True
        
        self.log_message("Yeni hesap bilgileri gerekiyor. Lütfen girin.", "USER_INPUT")
        dialog = CredentialsDialog(self, title="Hesap Değişikliği/Sorunu")
        self.credentials_dialog_open = False

        if dialog.result:
            new_creds = dialog.result
            if not new_creds['username'] or not new_creds['password']:
                messagebox.showerror("Eksik Bilgi", "Kullanıcı adı ve şifre gereklidir. İşlem duraklatıldı.")
                self.update_status("Hesap bilgisi bekleniyor (Eksik Giriş)")
                return

            if self.scraper and self.scraper.loop and not self.scraper.loop.is_closed():
                asyncio.run_coroutine_threadsafe(
                    self.scraper.switch_account_and_resume(new_creds, resume_state_info),
                    self.scraper.loop
                )
            else:
                self.log_message("Scraper loop aktif değil. Hesap değiştirilemiyor.", "ERROR")
        else: 
            self.log_message("Hesap değişikliği iptal edildi. Scraping duraklatıldı.", "WARN")
            self.update_status("Hesap bilgisi bekleniyor (İptal Edildi)")

    def toggle_pause_resume(self):
        if not self.scraper or not self.scraper.is_running: return
        if self.scraper.is_paused:
            self.scraper.resume_scraping()
            self.pause_resume_button.config(text="Duraklat")
        else:
            self.scraper.pause_scraping()
            self.pause_resume_button.config(text="Devam Ettir")
        self.save_button.config(state=tk.NORMAL if self.scraper.is_paused else tk.DISABLED)


    def handle_switch_account_button(self):
        if not self.scraper : 
            self.log_message("Scraper aktif değil.", "WARN")
            return
        
        current_state_to_resume = self.scraper.current_task_state
        if self.scraper.is_running and not self.scraper.is_paused:
            self.scraper.pause_scraping() 
            self.pause_resume_button.config(text="Devam Ettir")
            self.save_button.config(state=tk.NORMAL)

        self.handle_request_new_credentials(current_state_to_resume)

    def handle_save_button(self):
        if self.scraper:
            self.scraper.save_current_data()
    
    def handle_stop_button(self):
        if self.scraper and self.scraper.is_running:
            if messagebox.askyesno("Durdur ve Çık", "Scraping işlemini durdurup çıkmak istediğinize emin misiniz? Excel dosyası kaydedilecek."):
                self.scraper.stop_scraping()
        else: 
             self.destroy()


    def on_scraping_operation_finished(self):
        self.log_message("Scraping operasyonu sonlandı.", "INFO")
        self.update_gui_for_scraping_active(False)
        if hasattr(self, 'pause_resume_button'): self.pause_resume_button.config(text="Duraklat")
        if hasattr(self, 'status_label'): self.update_status("Bitti/Durduruldu")
        
        if messagebox.askyesno("İşlem Bitti", "Scraping tamamlandı veya durduruldu.\nAna menüye dönmek ister misiniz? (Hayır = Uygulamayı Kapat)"):
            if hasattr(self, 'main_app_frame') and self.main_app_frame.winfo_exists():
                 self.main_app_frame.pack_forget()
            self.init_query_builder_ui()
        else:
            self.destroy()


    def update_gui_for_scraping_active(self, is_active):
        state = tk.NORMAL if is_active else tk.DISABLED
        if hasattr(self, 'pause_resume_button'): self.pause_resume_button.config(state=state)
        if hasattr(self, 'switch_account_button'): self.switch_account_button.config(state=state)
        
        save_btn_state = tk.DISABLED
        if not is_active: 
            save_btn_state = tk.NORMAL
        elif self.scraper and self.scraper.is_paused: 
            save_btn_state = tk.NORMAL

        if hasattr(self, 'save_button'): self.save_button.config(state=save_btn_state)
        if hasattr(self, 'stop_button'): self.stop_button.config(text="Durdur ve Çık" if is_active else "Çıkış")


    def on_closing(self):
        if self.scraper and self.scraper.is_running:
            if messagebox.askyesno("Çıkış", "Scraping devam ediyor. Çıkmak istediğinize emin misiniz? Excel dosyası kaydedilecek."):
                self.scraper.stop_scraping() 
                self.destroy() 
            else:
                return 
        else:
            self.destroy()


if __name__ == '__main__':
    app = App()
    app.protocol("WM_DELETE_WINDOW", app.on_closing)
    app.mainloop()
